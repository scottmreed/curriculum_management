from openpyxl import Workbook
import pandas as pd
import os
import time

from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import decimal
from pathlib import Path
from datetime import date
import glob
import datetime
import numpy as np


def format_number(num):
    try:
        dec = decimal.Decimal(num)
    except:
        return num
    val = dec
    return val

def latest_edited_file_glob(pathToDir, **kwargs):
    pathTest = os.path.dirname(__file__)

    if pathToDir == 'ori':
        ori_file_path = os.path.join(pathTest, 'data', 'ori', 'CLSSCHED.CSV')
        if not os.path.isfile(ori_file_path):
            print('ori is empty. assume you want to compare to most recent color diff')
            latest_file = 'empty'
        else:
            list_of_files = glob.glob(
                os.path.join(pathTest, "data", "ori", '*'))  # * means all if need specific format then *.csv
            latest_file = max(list_of_files, key=os.path.getctime)
            print("glob latest_file:" + latest_file)

    elif pathToDir == 'changed':
        list_of_files = glob.glob(
            os.path.join(pathTest, "data", "changed", '*'))  # * means all if need specific format then *.csv
        latest_file = max(list_of_files, key=os.path.getctime)
        print("glob latest_file:" + latest_file)

    else:
        list_of_files = glob.glob(
            os.path.join(pathTest, "color_diffs", "Spring 2022", '*'))  # * means all if need specific format then *.csv
        latest_file = max(list_of_files, key=os.path.getctime)
        print("glob latest_file:" + latest_file)

    return latest_file


def read_csv(filename):
    # latest_edited_file_glob('test') # to get date
    path = os.path.dirname(__file__)
    print('File name :    ', os.path.basename(__file__))
    print('Directory Name:     ', path)

    if filename == 'ori':
        #latest_edited_file = max([f for f in os.scandir(os.path.join(path, "data", "ori"))], key=lambda x: x.stat().st_mtime).name
        #pathFile = os.path.join(path, "data", "ori", latest_edited_file)
        pathFile = latest_edited_file_glob('ori')

        try:
            modified = os.path.getmtime(pathFile)
            year, month, day, hour = time.localtime(modified)[:-5]
            file_date = str(str(year) + '_' + str(month) + '_' + str(day))
            print("Original Date:", file_date)
            print('pathFile is:', pathFile)
        except:
            file_date = time.ctime(os.path.getmtime(pathFile))

    elif filename == 'new':
        pathFile = latest_edited_file_glob('changed')
        if pathFile == 'empty':
            print('')
        try:
            modified = os.path.getmtime(pathFile)
            year, month, day, hour = time.localtime(modified)[:-5]
            file_date = str(str(year)+'_'+str(month)+'_'+str(day))
            print ("New Date:", file_date)
            file_date = str(str(year) + '_' + str(month) + '_' + str(day))
            print("New Date:", file_date)
        except:
            file_date = time.ctime(os.path.getmtime(pathFile))
        print('pathFile is:', pathFile)

    df = pd.read_csv(pathFile, header=None)

    return df, file_date


def duplicate_entry_merger(df):
    df = df.applymap(str)
    df = df.groupby([df.columns[0], df.columns[7]]).agg(" ; ".join).reset_index()
    return df


def time_conflict(df):

    filtered_columns_TC = [0, 4, 7, 15, 16, 17, 18, 19, 21, 23, 25, 27, 34, 35, 36, 38, 51, 53, 57, 58, 59, 66, 67, 68,
                           69, 70, 71, 72, 73, 74, 75, 76, 81, 82, 83, 85, 90, 91, 97, 108, 110, 112]

    len_array = len(df.columns)

    if len_array > 140:  # avoids applying filters to short files
        df = df.iloc[:, filtered_columns_TC]

    for i in range(df.shape[0]):
        for j in range(df.shape[1]):
            df.iloc[i, j] = format_number(df.iloc[i, j])

    df = df.replace("TA", np.nan)
    # df = df.dropna(axis=0)
    df['conflict_group'] = ''

    if len_array > 140:
        instructor_col = 34
        load_col = 39
        type_col = 8
    else:
        instructor_col = 13
        load_col = 15
        type_col = 6
    print(df.iloc[:, instructor_col])
    instructor_list = df.iloc[:, instructor_col].unique().tolist()
    # instructor_list = list(filter(str.strip, instructor_list))
    instructor_load = {"Ins":"load"}

    print(instructor_list)
    print(len(instructor_list))

    conflict_df_new = pd.DataFrame()

    for i in range(len(instructor_list)):
        temp_df_all = df.loc[df.iloc[:, instructor_col] == instructor_list[i]]
        load = temp_df_all.values[:, load_col:load_col+1]
        course_list = list(temp_df_all.values[:, 3])
        course_delta = [j - i for i, j in zip(course_list[:-1], course_list[1:])]
        load_sum = float(sum(np.sum(load, 0)))

        lower_lab = [2038, 2068, 3418, 3428, 3488]
        lab_count = len([x for x in course_list if x in lower_lab])
        load_sum -= lab_count
        load_sum += ((lab_count) * 0.2)
        lab_type_count = len([x for x in np.unique(course_list) if x in lower_lab])
        load_sum += lab_type_count

        if 2068 in course_list:
            load_sum -= len([x for x in course_list if x == 2068])

        for x in range(len(load)):
            course_num = (temp_df_all.values[x, 3])
            course_type = (temp_df_all.values[x, type_col])
            course_credit = (temp_df_all.values[x, load_col])

            if course_num > 5000 and 1000 in course_delta:
                load_sum -= course_credit

            if course_num == 3498:
                if str(course_type).startswith('Lab') or str(course_type).startswith('Main Lab'):
                    load_sum += 1

            if course_num == 2039 or course_num == 2069:
                load_sum -= .75

            if course_num == 3018 or course_num == 3118 or \
                    course_num == 4128 or course_num == 4518 or\
                    course_num == 4538 or course_num == 4548 or course_num == 4828:
                if str(course_type).startswith('Lab') or str(course_type).startswith('Main Lab'):
                    load_sum += .25
        instructor_load[str(instructor_list[i])] = load_sum
        print('instructor ', instructor_list[i], 'has load ', load_sum, 'from ', len(load), 'courses')
        print(temp_df_all)
        if len_array < 140:
            print('not checking for time conflicts')
        else:
            for day in range(7):
                temp_df = temp_df_all.loc[df.iloc[:, 23 + day] == 'Y']

                if len(temp_df) > 1:
                    for i in range(len(temp_df)):
                        st1 = datetime.datetime.strptime(temp_df.iloc[i, 21], '%I:%M:%S.%f_%p')
                        et1 = datetime.datetime.strptime(temp_df.iloc[i, 22], '%I:%M:%S.%f_%p')
                        duplicate_avoider = 0

                        for j in range(len(temp_df)):
                            if i != j:
                                st2 = datetime.datetime.strptime(temp_df.iloc[j, 21], '%I:%M:%S.%f_%p')
                                et2 = datetime.datetime.strptime(temp_df.iloc[j, 22], '%I:%M:%S.%f_%p')

                                if st1 < st2 < et1 or st1 < et2 < et1:
                                    print('CONFLICT DETECTED')
                                    print(temp_df.iloc[i, 21] + '//' + temp_df.iloc[i, 22])
                                    print('conflicts with')
                                    print(temp_df.iloc[j, 21] + '//' + temp_df.iloc[j, 22])

                                    temp_df.iloc[i, 41] = i
                                    temp_df.iloc[j, 41] = i

                                    if duplicate_avoider == 0:
                                        conflict_df_new = conflict_df_new.append(temp_df.iloc[i], ignore_index=True)
                                        duplicate_avoider = 1

                                    conflict_df_new = conflict_df_new.append(temp_df.iloc[j], ignore_index=True)

    conflict_df_new.to_csv('conflict_detected.csv', sep=',', index=False, header=False)
    return instructor_load


if __name__ == '__main__':
    today = date.today()

    df_modified_input_file, new_date = read_csv('new')
    df_modified_file = df_modified_input_file#.loc[df_modified_input_file.iloc[:, 14] == "CHEM"]

    for i in range(df_modified_file.shape[0]):
        for j in range(df_modified_file.shape[1]):
            df_modified_file.iloc[i, j] = format_number(df_modified_file.iloc[i, j])

    filtered_columns = [0, 4, 7, 15, 16, 17, 18, 19, 21, 23, 25, 27, 34, 35, 36, 38, 51, 53, 57,
                        66, 67, 81, 82, 83, 90, 91, 97, 106, 110, 112]
    df_filtered_mod = df_modified_file.iloc[:]

    if len(df_modified_file.columns) > 140: # avoids applying filters to short files
        instructor_load = time_conflict(df_modified_file)
        df_filtered_mod = df_modified_file.iloc[:, filtered_columns]
        modified_term = df_filtered_mod.iloc[0, 1]
    else:
        instructor_load = time_conflict(df_modified_file)
        modified_term = df_filtered_mod.iloc[0, 0]
    print('modified_term is:', modified_term)
    pd.DataFrame(instructor_load.items()).to_csv(f'{modified_term}_instructor_load.csv', sep=',',
                                                                   index=False, header=False)
    pathTest = os.path.dirname(__file__)
    pathFile = latest_edited_file_glob('ori')
    if pathFile == 'empty':
        pathFile = os.path.join(pathTest, "color_diffs", f"{modified_term}")
        list_of_files = glob.glob(
            os.path.join(pathFile, '*.CSV'))  # * means all if need specific format then *.csv
        if list_of_files:
            latest_file = max(list_of_files, key=os.path.getmtime)
            pathFile = latest_file
            print("glob latest_file:" + latest_file)
            modified = os.path.getmtime(pathFile)
            year, month, day, hour = time.localtime(modified)[:-5]
            ori_date = str(str(year) + '_' + str(month) + '_' + str(day))
        else:
            Path("color_diffs/" + modified_term).mkdir(parents=True, exist_ok=True)
        # modified = os.path.getmtime(pathFile)
        # year, month, day, hour = time.localtime(modified)[:-5]
        # ori_date = str(str(year) + '_' + str(month) + '_' + str(day))

    # ori_file_path = os.path.join(pathTest, 'data', 'ori', 'CLASSCHED.csv')
    # if not os.path.isfile(ori_file_path):
    #     print('ori is empty. assume you want to compare to most recent color diff')
    #     pathFile = os.path.join(pathTest, "color_diffs", f"{modified_term}", '*')

        # ori_date = time.ctime(os.path.getmtime(pathFile))
    else:
        modified = os.path.getmtime(pathFile)
        year, month, day, hour = time.localtime(modified)[:-5]
        ori_date = str(str(year) + '_' + str(month) + '_' + str(day))
        print("Original Date:", ori_date)
        print('pathFile is:', pathFile)

    df_original_file = pd.read_csv(pathFile, header=None)
    dept_col = 14
    if len(df_original_file.columns) < 140:
        dept_col = 2
    # df_original_file, ori_date = read_csv('ori')
    df_original_file = df_original_file.loc[df_original_file.iloc[:, dept_col] == "CHEM"]

    for i in range(df_original_file.shape[0]):
        for j in range(df_original_file.shape[1]):
            df_original_file.iloc[i, j] = format_number(df_original_file.iloc[i, j])


    df_filtered_ori = df_original_file.iloc[:]

    if len(df_original_file.columns) > 140:  # avoids applying filters to short files
       df_filtered_ori = df_original_file.iloc[:, filtered_columns]


    # print('83 ', df_filtered_mod[83])
    # print('67 ', df_filtered_mod[67:68])
    #
    # df_time = df_filtered_mod.iloc[:,[19,20]]
    # time_start = (df_filtered_mod.iloc[:,19])
    # time_stop = (df_filtered_mod.iloc[:, 20])
    # date_time_start = datetime.datetime.strptime(time_start[1], '%H:%M:%S.%f_%p')
    # date_time_end = datetime.datetime.strptime(time_stop[1], '%H:%M:%S.%f_%p')
    # time_range = date_time_end - date_time_start
    # print('df_time ', time_range)

    # df_instructor = df_filtered_mod.iloc[:,[3,6,8,17,23]]
    # print(df_instructor.shape)
    # print('df_instructor ', df_instructor)
    # ff = pd.pivot_table(df_instructor, values=['83'], index=['53', '83'], columns=['53'], aggfunc=np.sum, fill_value=0)


    print('term' + modified_term)
    Path("color_diffs/" + modified_term).mkdir(parents=True, exist_ok=True)

    df_filtered_ori = duplicate_entry_merger(df_filtered_ori)
    df_filtered_mod = duplicate_entry_merger(df_filtered_mod)

    short_columns = [4, 7, 14, 15, 16, 18, 21, 35, 36, 65, 66, 67, 81, 83, 97, 106]
    try:
        df_modified_input_file_short = df_modified_input_file.iloc[:, short_columns]
    except:
        df_modified_input_file_short = df_modified_input_file

    df_modified_input_file_short.to_csv(f'color_diffs/{modified_term}/{new_date}_short_out.csv', sep=',', index=False, header=False)

    df_changes = pd.concat([df_filtered_ori.assign(type='original'), df_filtered_mod.assign(type='modified')])
    df_changes = df_changes.drop_duplicates(keep=False, subset=df_changes.columns.difference(['type']))
    # df_changes = df_changes.drop_duplicates(keep=False, subset=df_changes.columns.difference([83]))#83=Inst
    try:
       df_modified_file_short = df_modified_file.iloc[:, short_columns]
    except:
        df_modified_file_short = df_modified_file

    if len(df_modified_file.columns) > 140:
        df_modified_input_file.to_csv(f'color_diffs/{modified_term}/{new_date}_full_out.csv', sep=',', index=False,
                                      header=False)
        df_modified_file.to_csv(f'color_diffs/{modified_term}/{ori_date}_{new_date}_changed_full.csv', sep=',', index=False, header=False)
        df_modified_file_short = df_modified_file_short.drop_duplicates(keep='first', subset=[4, 7, 14, 15, 16, 18, 21])  # works sometimes

    else:
        df_modified_file_short = df_modified_file_short.drop_duplicates(keep='first', subset=[1,2,3,4,5,6,7,8,9])  # works sometimes

    df_modified_file_short.to_csv(f'color_diffs/{modified_term}/{ori_date}_{new_date}_changed_short.csv', sep=',', index=False, header=False)

    df_changes_arranged = df_changes

    df_changes_arranged = df_changes_arranged.sort_values(
        by=[df_changes_arranged.columns[0], df_changes_arranged.columns[7]], ascending=True) #arrange by coursenum and class ID

    j = 0
    i = 0
    c = 0
    rows_count = df_changes_arranged.shape[0]
    cols_count = df_changes_arranged.shape[1]

    changed_id = []
    unique_id = []

    unique_id_from_ori = []
    unique_id_from_mod = []

    red_color = 'background-color: red'
    green_color = 'background-color: lightgreen'

    for i in range(rows_count - 1):
        if ((df_changes_arranged.iloc[i, 0] == df_changes_arranged.iloc[i + 1, 0]) and (
                df_changes_arranged.iloc[i, 7] == df_changes_arranged.iloc[i + 1, 7])):
            j = 0

            for c in range(cols_count):
                if c < 29:
                    if df_changes_arranged.iloc[i, c] != df_changes_arranged.iloc[i + 1, c]:
                        changed_id.append([i, c])

        else:
            j = j + 1
            if j >= 2:
                unique_id.append([i + 1, 1])
                j = 0

            elif i == rows_count - 2:
                unique_id.append([i + 2, 1])
                j = 0

            else:
                'a'

    wb = Workbook()
    ws = wb.active

    for r in dataframe_to_rows(df_changes_arranged, index=False, header=False):
        ws.append(r)

    fill_pattern_modified_content = PatternFill(patternType='solid', fgColor='ffaba6')
    fill_pattern_modified_content_darker = PatternFill(patternType='solid', fgColor='f03629')
    fill_pattern_unique_content = PatternFill(patternType='solid', fgColor='61bdff')

    fill_pattern_from_original = PatternFill(patternType='solid', fgColor='FFA500')
    fill_pattern_from_modified = PatternFill(patternType='solid', fgColor='66CDAA')

    for l in range(len(df_changes_arranged)):
        if ws.cell(l + 1, 30).value == "original":
            for p in range(cols_count):
                ws.cell(l + 1, p + 1).fill = fill_pattern_from_original
        elif ws.cell(l + 1, 30).value == "modified":
            for p in range(cols_count):
                ws.cell(l + 1, p + 1).fill = fill_pattern_from_modified

    for i in range(len(changed_id)):
        ws.cell(changed_id[i][0] + 1, changed_id[i][1] + 1).fill = fill_pattern_modified_content
        ws.cell(changed_id[i][0] + 2, changed_id[i][1] + 1).fill = fill_pattern_modified_content_darker

    wb.save("color_diffs/" + modified_term + "/" + today.isoformat() + "highlighted_changes.xlsx")
    try:
        if len(df_modified_file.columns) < 140:
            wb2 = Workbook()
            ws1 = wb2.active
            ws1.title = "Load"
            ws1.append(["name", "Load"])
            for i, (k, v) in enumerate(instructor_load.items()):

                ws1.cell(i+1, 1).value = k
                ws1.cell(i+1, 2).value = v
            # wb2.save("color_diffs/" + modified_term + "/" + today.isoformat() + "_load_output.xlsx")

    except:
        'a'